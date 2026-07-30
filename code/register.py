"""Reading the selected-acts register, in the order the extraction stages want it.

Every extraction stage needs the same two things from `hcj_acts_selected.xlsx`:
which acts belong to it, and in what order to work through them. Both were
previously answered three times over, and the second one was answered wrongly —
the stages globbed the Markdown directory and sorted it, which sorts by *file
name*, i.e. by act serial as a string. `--limit 50` therefore extracted acts
1004…1049, an arbitrary slice of 2025, rather than the 50 most recent acts.

Order here is **newest act first**, by `Дата прийняття` and then by serial within
a day. That is the order a corpus is worth building in: the newest acts are the
ones a reader asks about, and a partial extraction that stops halfway should
leave the recent end covered rather than a random middle.

Rows without a downloaded file are dropped — there is no text to extract from —
and rows whose date does not parse sort last rather than being discarded, so a
malformed date shows up as an act extracted late instead of an act silently
missing.
"""

from datetime import date, datetime, time, timedelta
from pathlib import Path

import openpyxl

import act_numbers

FILE_COL = "Локальний файл"
DATE_COL = "Дата прийняття"
NUMBER_COL = "Номер"
KIND_COL = "Вид документу"
SEEN_COL = "Вперше побачено"

_EPOCH = datetime.min  # unparseable dates sort last, they are not dropped


def _as_date(value) -> datetime:
    if hasattr(value, "year"):  # openpyxl hands back a datetime for date-typed cells
        return datetime(value.year, value.month, value.day)
    try:
        return datetime.strptime(str(value).strip(), "%d.%m.%Y")
    except (ValueError, TypeError):
        return _EPOCH


def _serial(number) -> int:
    parsed = act_numbers.parse_act_number(number)
    return parsed["serial"] if parsed else 0


def stem_of(record: dict) -> str:
    """The Markdown/raw file stem for a register row (`1503_22.07.2026`)."""
    return str(record[FILE_COL]).rsplit(".", 1)[0]


def records(register_path: Path) -> list[dict]:
    """Register rows that have a downloaded file, newest act first."""
    ws = openpyxl.load_workbook(register_path, read_only=True).active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = [dict(zip(header, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    rows = [r for r in rows if r.get(FILE_COL)]
    rows.sort(key=lambda r: (_as_date(r.get(DATE_COL)), _serial(r.get(NUMBER_COL))), reverse=True)
    return rows


def stems(register_path: Path, keep) -> list[str]:
    """File stems of the rows `keep(record)` accepts, newest act first.

    Returns a list, not a set: the order is the point. Callers map it onto the
    Markdown directory so that `--limit N` means "the N most recent acts of this
    kind that are not extracted yet".
    """
    return [stem_of(r) for r in records(register_path) if keep(r)]


def seen_within(days: int):
    """Predicate accepting acts the register first discovered in the last `days`.

    This is what makes the nightly job safe to leave unattended. «Вперше
    побачено» is stamped only for acts a scrape actually found for the first
    time — seeded history carries an empty stamp on purpose (11_scrape_register
    calls it "an honest change log"), and re-reading an act never re-stamps it.
    So this selects last night's arrivals and *excludes the backlog*, which is
    the difference between a nightly bill of cents and one of tens of dollars.

    An act with no stamp is never "new". That is the safe direction to fail: a
    missed act shows up as a gap on the site, where filtering the wrong way
    would silently spend the corpus budget.
    """
    cutoff = datetime.combine(date.today() - timedelta(days=days), time.min)

    def keep(record: dict) -> bool:
        stamp = record.get(SEEN_COL)
        return bool(stamp) and _as_date(stamp) >= cutoff

    return keep


def both(*predicates):
    """A predicate accepting records that every one of `predicates` accepts."""
    return lambda record: all(p(record) for p in predicates)


def markdown_files(markdown_dir: Path, ordered_stems: list[str],
                   only: list[str] | None = None) -> list[Path]:
    """Markdown paths for `ordered_stems`, keeping that order and skipping gaps.

    `only` narrows the result to named stems. Newest-first is the right default,
    but it cannot produce a linked case: an ухвала and the рішення it opens are
    six to eighteen months apart, so any recent window holds one end or the other
    and never both. Naming acts explicitly is how a whole case gets extracted.
    """
    available = {p.stem: p for p in markdown_dir.glob("*.md")}
    stems = ordered_stems if only is None else [s for s in ordered_stems if s in set(only)]
    return [available[stem] for stem in stems if stem in available]
