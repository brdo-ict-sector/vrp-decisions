"""
Extract the «номер дисциплінарної скарги» — the HCJ incoming-correspondence number
under which a disciplinary complaint is registered — from the Markdown acts.

Why: a рішення and the ухвала that opened the case carry different act numbers
(e.g. 2551/2дп/15-25 and 1408/2дп/15-26) but quote the same complaint number
(А-1720/0/7-25). That number is the only key that joins the two document types
into one disciplinary case.

Anatomy of the number:   Л - 1720 / 0 / 7 - 25
                         │   │     │   │    └─ year the complaint was registered
                         │   │     │   └────── register index = who complained
                         │   │     └────────── sub-number (доповнення, пояснення…)
                         │   └──────────────── serial number in that register
                         └──────────────────── first letter of the complainant's
                                               surname (individuals only)

Register indexes seen in the corpus, and how this script treats them:

    7   complaint from an individual (letter-prefixed)      → always accepted
    13  complaint from a state body (НАЗК, НАБУ, прокуратура)→ always accepted
    8   general incoming mail; carries complaints from       → accepted only in a
        officials (court chairs, Ombudsman) but also           complaint context
        court letters, ВККС references, characteristics
    6   the judge's own file (пояснення, скарги судді on     → rejected
        a decision) — never a disciplinary complaint
    9, 15, 19, 24, 149, 166 …  outgoing acts, inspector      → rejected
        conclusions, other registers

Because a case file grows sub-numbers (Ч-444/0/7-25, Ч-444/16/7-25,
Ч-444/17/7-25 are one complaint), the script also emits a `case_key`
— `serial/index-year`, without the sub-number and without the letter — which is
what documents should actually be joined on. The letter is dropped from the key
because the serial is already unique within an index+year, and because document
conversion sometimes loses the letter (`М‑2522/45/7-21` → `2522/45/7-21`).

Usage:
    python 14_extract_complaint_numbers.py                    # report over all acts
    python 14_extract_complaint_numbers.py --years 2025 2026
    python 14_extract_complaint_numbers.py --csv out.csv      # one row per act
    python 14_extract_complaint_numbers.py --update-xlsx      # add columns to the register
    python 14_extract_complaint_numbers.py --show 1408_08.07.2026
                                                               # every match with context

Stage 12 rebuilds hcj_acts_selected.xlsx from scratch on every run, so
--update-xlsx has to run after it — see code/run_daily.sh.
"""

import argparse
import csv
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

# ── Config ──────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
MD_DIR = BASE_DIR / "data" / "acts" / "md"
SELECTED_XLSX = BASE_DIR / "data" / "register" / "hcj_acts_selected.xlsx"

# Columns --update-xlsx maintains on the selected-acts register.
XLSX_COLUMNS = ("Номер дисциплінарної скарги", "Ключ справи")

# Register indexes that are always a disciplinary complaint.
INDEX_ALWAYS = {"7", "13"}
# Register indexes that are a complaint only when the surrounding text says so.
INDEX_IF_CONTEXT = {"8"}

# How far around an index-8 number the complaint wording has to be. The window is
# asymmetric because the giveaway phrase almost always follows the number
# ("(вх. № 7854/0/8-25) надійшла дисциплінарна скарга"), while a wide window to
# the right starts catching the next paragraph — court letters and inspector
# conclusions quoted a sentence away from an unrelated mention of «скарга».
CONTEXT_BEFORE = 130
CONTEXT_AFTER = 70

# A number the context test rejected is still part of a complaint file if it sits
# right next to an accepted one — either listed in the same parenthetical
# ("(вх. № 5264/0/8-23, № 7965/1/8-25)") or as a sub-number of the same serial
# (доповнення 4126/1/8-24 to complaint 4126/0/8-24).
ADJACENCY_CHARS = 40

# ── Normalisation ───────────────────────────────────────────────────────────
# Docling emits several dash and space code points; the number must survive them.
DASHES = "‐‑‒–—―−－"
SPACES = "       "
_DASH_RE = re.compile(f"[{DASHES}]")
_SPACE_RE = re.compile(f"[{SPACES}]")

# Latin lookalikes that OCR/typing leaves in place of Cyrillic prefixes.
LATIN_TO_CYRILLIC = str.maketrans("ABCEHIKMOPTXY", "АВСЕНІКМОРТХУ")


def normalize(text: str) -> str:
    """Flatten a Markdown act to one line with canonical dashes and single spaces."""
    text = _DASH_RE.sub("-", _SPACE_RE.sub(" ", text))
    return re.sub(r"\s+", " ", text)


# ── The number ──────────────────────────────────────────────────────────────
# The letter prefix may be separated from the serial by a line break the
# converter turned into a space ("вх. № К- 1896/1/7-25"), or lose its hyphen
# entirely ("№№ Б 496/0/7-24"), so both separators are optional. The acts
# themselves also mistype the last separator as "/-" ("К-698/52/7/-25").
NUMBER_RE = re.compile(
    r"(?<![\w/-])"
    r"(?:(?P<letter>[А-ЯІЇЄҐA-Z]) ?-? ?)?"
    r"(?P<serial>\d{1,5}) ?/ ?(?P<sub>\d{1,3}) ?/ ?(?P<index>\d{1,3}) ?/? ?- ?(?P<year>\d{2})"
    r"(?!\d)"
)

# Strings shaped like a complaint number that the pattern deliberately refuses:
# a four-digit year (the acts occasionally double it — "К-698/52/7-2524"), or a
# letter prefix on a register that is not a complaint register. These are source
# typos rather than data, so they are reported instead of silently guessed at.
NEAR_MISS_RE = re.compile(
    r"(?<![\w/-])(?:[А-ЯІЇЄҐA-Z] ?-? ?)?\d{1,5} ?/ ?\d{1,3} ?/ ?\d{1,3} ?/? ?- ?\d{4}(?!\d)"
)

# Phrases that mark the number as belonging to a complaint rather than to a
# court letter or a reference document.
COMPLAINT_CONTEXT_RE = re.compile(
    r"дисциплінарн\w*\s+скарг"
    r"|скарг\w*\s+(?:на\s+ді|стосовно|щодо|голови|заступника|Уповноваженого|прокурор)"
    r"|за\s+скарг\w+"
    r"|надійшл\w+\s+(?:\S+\s+){0,4}скарг"
    r"|доповнення\s+до\s+(?:дисциплінарної\s+)?скарг",
    re.IGNORECASE,
)


class Match:
    """One accepted or rejected number occurrence."""

    def __init__(self, number, case_key, index, span, accepted, reason, context):
        self.number = number
        self.case_key = case_key
        self.index = index
        self.span = span  # (start, end) in the normalised text
        self.accepted = accepted
        self.reason = reason
        self.context = context


def find_numbers(text: str) -> list[Match]:
    """Find every вх-style number in an act and decide which are complaint numbers."""
    text = normalize(text)
    out = []
    for m in NUMBER_RE.finditer(text):
        letter, serial, sub, index, year = m.group("letter", "serial", "sub", "index", "year")
        if letter:
            letter = letter.translate(LATIN_TO_CYRILLIC)
        number = f"{letter}-{serial}/{sub}/{index}-{year}" if letter else f"{serial}/{sub}/{index}-{year}"
        case_key = f"{serial}/{index}-{year}"

        if index in INDEX_ALWAYS:
            accepted, reason = True, f"index {index}"
        elif index in INDEX_IF_CONTEXT:
            window = text[max(0, m.start() - CONTEXT_BEFORE) : m.end() + CONTEXT_AFTER]
            accepted = bool(COMPLAINT_CONTEXT_RE.search(window))
            reason = f"index {index} {'+ complaint context' if accepted else 'without complaint context'}"
        else:
            accepted, reason = False, f"index {index} is not a complaint register"

        context = text[max(0, m.start() - 110) : m.start()] + f"«{m.group(0)}»" + text[m.end() : m.end() + 110]
        out.append(Match(number, case_key, index, m.span(), accepted, reason, context))

    return _accept_neighbours(out)


def find_near_misses(text: str) -> list[str]:
    """Number-shaped strings the pattern refused — almost always a typo in the act."""
    text = normalize(text)
    return [re.sub(r"\s+", "", m.group(0)) for m in NEAR_MISS_RE.finditer(text)]


def _accept_neighbours(matches: list[Match]) -> list[Match]:
    """Second pass: pull in context-rejected numbers that belong to an accepted file.

    Only numbers from the context-gated registers are reconsidered — the rejected
    indexes (the judge's own file, outgoing acts, inspector conclusions) are
    rejected structurally and must stay out however close they sit.
    """
    accepted_keys = {m.case_key for m in matches if m.accepted}
    accepted_spans = [m.span for m in matches if m.accepted]
    for m in matches:
        if m.accepted or m.index not in INDEX_IF_CONTEXT:
            continue
        if m.case_key in accepted_keys:
            m.accepted, m.reason = True, f"index {m.index} + sub-number of an accepted complaint"
        elif any(s < m.span[1] + ADJACENCY_CHARS and m.span[0] - ADJACENCY_CHARS < e for s, e in accepted_spans):
            m.accepted, m.reason = True, f"index {m.index} + listed beside an accepted complaint"
    return matches


def act_year(stem: str) -> str:
    """'1408_08.07.2026' → '2026'."""
    return stem.rsplit(".", 1)[-1]


def collect(md_dir: Path, years: list[str] | None):
    """Return {stem: [Match, …]} for the acts in scope, in file-name order."""
    results = {}
    for f in sorted(md_dir.glob("*.md")):
        if years and act_year(f.stem) not in years:
            continue
        results[f.stem] = find_numbers(f.read_text(encoding="utf-8"))
    return results


def collect_near_misses(md_dir: Path, years: list[str] | None):
    """Return {stem: [malformed number, …]} for the acts in scope."""
    out = {}
    for f in sorted(md_dir.glob("*.md")):
        if years and act_year(f.stem) not in years:
            continue
        misses = find_near_misses(f.read_text(encoding="utf-8"))
        if misses:
            out[f.stem] = misses
    return out


def accepted_numbers(matches: list[Match]) -> list[str]:
    """Distinct accepted numbers, in order of first appearance."""
    return list(dict.fromkeys(m.number for m in matches if m.accepted))


def accepted_case_keys(matches: list[Match]) -> list[str]:
    return list(dict.fromkeys(m.case_key for m in matches if m.accepted))


# ── Reporting ───────────────────────────────────────────────────────────────
def report(results, near_misses=None):
    total = len(results)
    per_doc = {stem: accepted_numbers(ms) for stem, ms in results.items()}
    per_doc_keys = {stem: accepted_case_keys(ms) for stem, ms in results.items()}
    empty = [s for s, n in per_doc.items() if not n]

    print(f"📄 Acts scanned: {total}")
    print(f"   with ≥1 complaint number: {total - len(empty)}")
    print(f"   with none:                {len(empty)}")

    print("\n── Numbers per act ──")
    dist = Counter(len(n) for n in per_doc.values())
    for n in sorted(dist):
        print(f"   {n:>3} number(s): {dist[n]:>4} acts")

    print("\n── Distinct case keys after joining ──")
    by_key = defaultdict(list)
    for stem, keys in per_doc_keys.items():
        for k in keys:
            by_key[k].append(stem)
    multi = {k: v for k, v in by_key.items() if len(v) > 1}
    print(f"   distinct case keys:        {len(by_key)}")
    print(f"   keys shared by >1 act:     {len(multi)}")
    print(f"   acts joined into a case:   {len({s for v in multi.values() for s in v})}")

    print("\n── Acts with more than one complaint number (first 25) ──")
    many = sorted(((s, n) for s, n in per_doc.items() if len(n) > 1), key=lambda x: -len(x[1]))
    print(f"   {len(many)} such acts")
    for stem, nums in many[:25]:
        shown = ", ".join(nums[:8]) + (" …" if len(nums) > 8 else "")
        print(f"   {stem:<22} {len(nums):>2}  {shown}")

    print("\n── Acts with no complaint number (first 20 of %d) ──" % len(empty))
    for stem in empty[:20]:
        print(f"   {stem}")

    print("\n── Sample joined cases (first 10 with ≥2 acts) ──")
    for key, stems in sorted(multi.items(), key=lambda kv: -len(kv[1]))[:10]:
        print(f"   {key:<16} {len(stems):>2} acts: {', '.join(sorted(stems)[:6])}")

    print("\n── Register index seen (accepted / rejected occurrences) ──")
    idx = Counter()
    for ms in results.values():
        for m in ms:
            idx[(m.index, m.accepted)] += 1
    for index in sorted({i for i, _ in idx}, key=lambda i: -(idx[(i, True)] + idx[(i, False)])):
        print(f"   index {index:>3}: accepted {idx[(index, True)]:>5}   rejected {idx[(index, False)]:>5}")

    print("\n── Malformed numbers in the source (check by hand) ──")
    if near_misses:
        for stem, misses in near_misses.items():
            print(f"   {stem:<22} {', '.join(sorted(set(misses)))}")
    else:
        print("   none")


def show(results, stem):
    matches = results.get(stem)
    if matches is None:
        sys.exit(f"No such act in scope: {stem}")
    print(f"── {stem} ──")
    for m in matches:
        mark = "✅" if m.accepted else "  "
        print(f"{mark} {m.number:<20} case={m.case_key:<14} {m.reason}")
        print(f"     …{m.context}…\n")


def update_xlsx(results, path: Path):
    """Add/refresh the complaint-number columns on the selected-acts register.

    Rows are matched on «Локальний файл» — the same stem the whole pipeline keys
    on. The columns are appended if missing and overwritten if already there, so
    re-running after stage 12 has rebuilt the workbook restores them.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]

    targets = {}
    for name in XLSX_COLUMNS:
        if name in header:
            targets[name] = header.index(name) + 1
        else:
            targets[name] = len(header) + 1
            header.append(name)
            cell = ws.cell(row=1, column=targets[name], value=name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(targets[name])].width = 28

    try:
        stem_col = header.index("Локальний файл") + 1
    except ValueError:
        sys.exit(f"{path} has no «Локальний файл» column — is this the selected-acts register?")

    filled = 0
    for row in ws.iter_rows(min_row=2):
        local_file = row[stem_col - 1].value
        if not local_file:
            continue
        matches = results.get(str(local_file).rsplit(".", 1)[0])
        if matches is None:
            continue  # act not converted to Markdown yet — leave the cells alone
        values = ("; ".join(accepted_numbers(matches)), "; ".join(accepted_case_keys(matches)))
        for name, value in zip(XLSX_COLUMNS, values):
            cell = ws.cell(row=row[0].row, column=targets[name], value=value or None)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        filled += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    tmp = path.with_name(path.name + ".tmp")
    wb.save(tmp)
    os.replace(tmp, path)
    print(f"\n💾 {filled} rows updated in {path}")


def write_csv(results, path):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["act", "complaint_numbers", "case_keys", "n_numbers"])
        for stem, ms in results.items():
            nums, keys = accepted_numbers(ms), accepted_case_keys(ms)
            w.writerow([stem, "; ".join(nums), "; ".join(keys), len(nums)])
    print(f"\n💾 {len(results)} rows → {path}")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--md-dir", type=Path, default=MD_DIR)
    ap.add_argument("--years", nargs="*", help="keep only acts adopted in these years, e.g. --years 2025 2026")
    ap.add_argument("--csv", type=Path, help="write one row per act to this CSV")
    ap.add_argument("--update-xlsx", nargs="?", const=SELECTED_XLSX, type=Path, metavar="XLSX",
                    help="add the complaint-number columns to the selected-acts register")
    ap.add_argument("--show", metavar="ACT", help="print every match in one act with its context")
    args = ap.parse_args()

    results = collect(args.md_dir, args.years)
    if args.show:
        show(results, args.show)
        return
    report(results, collect_near_misses(args.md_dir, args.years))
    if args.csv:
        write_csv(results, args.csv)
    if args.update_xlsx:
        update_xlsx(results, args.update_xlsx)


if __name__ == "__main__":
    main()
