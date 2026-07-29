"""
Export structured decisions from SQLite to docs/decisions.json for the web app,
joining each record to its entry in the scraped register.

The register (stage 12's selection) carries the official number, the document
type and the public ВРП URL for every act, keyed by the same
"<номер>_<дата>" filename the pipeline uses throughout. The round-2 web-links
spreadsheet is kept only as a fallback for records predating the register.

`doc_type` matters on the site: a «Ухвала про відкриття дисциплінарної справи»
opens a case and by its nature carries no sanction and no ВРП review stage —
without the type those empty fields read as failed extractions.

Writing to `docs/decisions.json` publishes: that file is what GitHub Pages serves.
`--output` exists so a batch can be reviewed in the UI first — export to a scratch
copy, look at it, and only then overwrite what the site serves.

Usage:
    python 32_export_to_json.py
    python 32_export_to_json.py --output /tmp/preview/decisions.json
"""

import argparse
import json
import sqlite3
from pathlib import Path

import openpyxl

import act_numbers
import joins

BASE_DIR = Path(__file__).resolve().parent.parent
DB_PATH = BASE_DIR / "data" / "decisions.db"
SELECTED_XLSX = BASE_DIR / "data" / "register" / "hcj_acts_selected.xlsx"
LINKS_XLSX = BASE_DIR / "data" / "reference" / "decisions to web-links.xlsx"
DOCS_DIR = BASE_DIR / "docs"
OUTPUT_JSON = DOCS_DIR / "decisions.json"


def load_register() -> dict[str, dict]:
    """Map decision filename stem -> {decision_num, doc_type, url}."""
    entries: dict[str, dict] = {}
    if not SELECTED_XLSX.exists():
        print(f"WARN: register selection not found: {SELECTED_XLSX}")
        return entries

    ws = openpyxl.load_workbook(SELECTED_XLSX, read_only=True).active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    col = {name: i for i, name in enumerate(header)}
    needed = ("Номер", "Вид документу", "doc_url", "Локальний файл")
    if any(name not in col for name in needed):
        print(f"WARN: unexpected columns in {SELECTED_XLSX.name}: {header}")
        return entries

    for row in rows:
        def val(name: str) -> str:
            i = col[name]
            return "" if i >= len(row) or row[i] is None else str(row[i]).strip()

        stem = Path(val("Локальний файл")).stem
        if not stem:
            continue
        entries[stem] = {
            "decision_num": val("Номер") or None,
            "doc_type": val("Вид документу") or None,
            "url": val("doc_url") or None,
        }
    return entries


def load_links() -> dict[str, dict]:
    """Fallback map: leading decision number -> {decision_num, url}."""
    links: dict[str, dict] = {}
    if not LINKS_XLSX.exists():
        return links
    wb = openpyxl.load_workbook(LINKS_XLSX, read_only=True)
    ws = wb.active
    for num, url in list(ws.iter_rows(values_only=True))[1:]:
        if not num:
            continue
        lead = str(num).split("/", 1)[0].strip()
        links[lead] = {"decision_num": str(num).strip(), "url": str(url).strip() if url else None}
    return links


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--output", type=Path, default=OUTPUT_JSON,
                    help="where to write the JSON (default: the published docs/decisions.json)")
    args = ap.parse_args()

    if not DB_PATH.exists():
        raise SystemExit(f"Database not found: {DB_PATH}\nRun 21_extract_decisions.py first.")

    register = load_register()
    links = load_links()

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT filename, data, processed_at FROM decisions "
            "WHERE data IS NOT NULL ORDER BY filename"
        ).fetchall()
        # Rulings and reviews are never exported in their own right: an ухвала that
        # is not attached to a decision has no card of its own on the site, because
        # "a case was opened" is not yet an answer to anything.
        rulings = {r["filename"]: json.loads(r["data"])
                   for r in conn.execute("SELECT filename, data FROM rulings WHERE data IS NOT NULL")}
        reviews = {r["filename"]: json.loads(r["data"])
                   for r in conn.execute("SELECT filename, data FROM reviews WHERE data IS NOT NULL")}

    def stage_grounds(qual: dict, stage: str) -> list:
        s = (qual or {}).get(stage)
        return (s or {}).get("grounds", []) if isinstance(s, dict) else []

    def flatten_judges(data: dict) -> list[dict]:
        """One list of judges, whatever shape the record was extracted in.

        Records written before the per-judge schema carry `judge_name` / `court`
        with qualification, conduct and sanction at the top level. They are not
        re-extracted automatically (the stage skips rows that already have data),
        so both shapes coexist until the corpus run replaces them. Reading them
        here keeps the published dataset uniform.
        """
        if isinstance(data.get("judges"), list):
            return data["judges"]
        if not data.get("judge_name"):
            return []
        sanction = data.get("sanction") or {}
        sanction_type = data.get("sanction_type") or {}
        return [{
            "name": data.get("judge_name"),
            "court": data.get("court"),
            "position": None,
            "qualification": data.get("qualification") or {},
            "conduct": data.get("conduct") or {},
            # Legacy records kept sanction per stage; the ВРП slot was never filled
            # because a review is a separate act, so the ДП value is the sanction.
            "sanction": sanction.get("dp") if isinstance(sanction, dict) else sanction,
            "sanction_type": (sanction_type.get("dp")
                              if isinstance(sanction_type, dict) else sanction_type),
            "outcome": None,
        }]

    decisions = []
    for r in rows:
        data = json.loads(r["data"])
        filename = r["filename"]
        lead = filename.split("_", 1)[0]
        # The register is keyed by the exact stem; the spreadsheet only by the
        # leading number, so it is consulted only where the register has no row.
        # (Not Path().stem — "1004_14.05.2025" would lose its ".2025".)
        stem = filename[:-3] if filename.endswith(".md") else filename
        link = register.get(stem) or links.get(lead, {})
        # Date from the decision, falling back to the filename suffix (983_12.05.2025).
        date = data.get("date") or (filename.split("_", 1)[1] if "_" in filename else None)

        judges = flatten_judges(data)
        lead_judge = judges[0] if judges else {}

        # ── Related acts ────────────────────────────────────────────────────
        ruling_hit = joins.find_ruling(data, filename, rulings)
        review_hit = joins.find_review(data, reviews)

        def related_act(hit: dict, num_key: str) -> dict | None:
            """Act-level facts about a related act, plus how the link was made."""
            if not hit:
                return None
            rec, rel_link = hit["record"], register.get(hit["filename"], {})
            return {
                "filename": hit["filename"],
                "num": rec.get(num_key) or rel_link.get("decision_num"),
                "date": rec.get("date"),
                "url": rel_link.get("url"),
                # R7: a provisional link must be visibly provisional.
                "matched_on": hit["matched_on"],
                "method": hit["method"],
                "other_candidates": hit.get("other_candidates", 0),
            }

        ruling_out = related_act(ruling_hit, "ruling_num")
        if ruling_out:
            rec = ruling_hit["record"]
            ruling_out.update({
                "complaint_number": rec.get("complaint_number"),
                "complaint_date": rec.get("complaint_date"),
                "complainant_name": rec.get("complainant_name"),
                "complainant_type": rec.get("complainant_type"),
                "inspector": rec.get("inspector"),
                "inspector_proposal": rec.get("inspector_proposal"),
                "facts": (rec.get("summary") or {}).get("facts"),
            })
        review_out = related_act(review_hit, "review_num")
        if review_out:
            rec = review_hit["record"]
            review_out.update({
                "appellant_type": rec.get("appellant_type"),
                "appellant_name": rec.get("appellant_name"),
                "essence": (rec.get("summary") or {}).get("essence"),
            })

        # Per judge, from that judge's OWN entry in the related act — an ухвала
        # opens different grounds against different judges, so taking the first
        # entry would attribute one judge's grounds to another.
        for j in judges:
            rj = joins.judge_view(ruling_hit["record"], j.get("name")) if ruling_hit else None
            j["ruling"] = {
                "grounds": (rj.get("grounds") or {}),
                "outcome": rj.get("outcome"),
                "judge_position": rj.get("judge_position"),
            } if rj else None
            vj = joins.judge_view(review_hit["record"], j.get("name")) if review_hit else None
            j["review"] = {
                "review_outcome": vj.get("review_outcome"),
                "qualification": vj.get("qualification") or {},
                # R5: where a review exists, the sanction in force is the review's.
                "sanction": vj.get("sanction"),
                "sanction_type": vj.get("sanction_type"),
            } if vj else None
        qual = lead_judge.get("qualification") or {}
        # Flat ground list for the facet filter: the palate's qualification, else
        # what the complaint alleged.
        grounds = stage_grounds(qual, "dp") or stage_grounds(qual, "complaint")

        decisions.append({
            "filename": filename,
            "date": date,
            "decision_num": data.get("decision_num") or link.get("decision_num"),
            "short_name": data.get("short_name"),
            "doc_type": link.get("doc_type"),
            "url": link.get("url"),
            # Flat lead-judge fields keep the current single-page app working while
            # the per-judge view is built; `judges` carries the full truth.
            "judge_name": lead_judge.get("name"),
            "court": lead_judge.get("court"),
            "judges": judges,
            # Every judge and court the act names, so the facets can filter on a
            # co-defendant rather than only on whoever happens to be listed first.
            "judge_names": [j["name"] for j in judges if j.get("name")],
            "courts": sorted({j["court"] for j in judges if j.get("court")}),
            # The two related acts, or null. An ухвала with no decision is not
            # exported at all — it appears here or nowhere.
            "ruling": ruling_out,
            "review": review_out,
            # Stated by the act number, not guessed by the model.
            "chamber": act_numbers.chamber_of(link.get("decision_num")) or data.get("chamber"),
            "art106_grounds": grounds,
            "qualification": qual,
            "conduct": lead_judge.get("conduct") or {},
            "sanction": lead_judge.get("sanction"),
            "sanction_type": lead_judge.get("sanction_type"),
            "summary": data.get("summary") or {},
            "processed_at": r["processed_at"],
        })

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(decisions, ensure_ascii=False, indent=2), encoding="utf-8")
    linked = sum(1 for d in decisions if d["url"])
    with_ruling = sum(1 for d in decisions if d["ruling"])
    with_review = sum(1 for d in decisions if d["review"])
    print(f"Exported {len(decisions)} decisions ({linked} with source links) → {args.output}")
    # R8: silence is indistinguishable from a bug, so the joins are always counted.
    print(f"  ухвала attached: {with_ruling}/{len(decisions)}"
          f"   ВРП перегляд attached: {with_review}/{len(decisions)}")
    orphan_rulings = len(rulings) - len({d["ruling"]["filename"] for d in decisions if d["ruling"]})
    print(f"  ухвал extracted but not attached to any decision (not exported): {orphan_rulings}")


if __name__ == "__main__":
    main()
