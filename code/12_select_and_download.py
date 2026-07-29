#!/usr/bin/env python3
"""
Stage 12: pick the disciplinary acts out of the register and fetch their files.

Selection (applied to data/register/hcj_acts.xlsx, produced by stage 00):

    «Ознака до документа» = «Результати розгляду питань щодо притягнення
    суддів до дисциплінарної відповідальності»
      → keep every «Рішення»
      → keep a «Ухвала» only if its «Назва документу» mentions «відкриття
        дисциплінарної справи» (that is the act opening a case; the far more
        numerous «Про відмову у *відкритті* дисциплінарної справи» and the
        procedural ухвали are deliberately left out)

The surviving rows are written to data/register/hcj_acts_selected.xlsx so the
selection can be inspected, and their «Файл акта» is downloaded into
data/acts/raw/ as «<номер>_<дата>.<ext>» — the same stem the whole pipeline
keys on, all the way to the record in SQLite.

Each document is fetched once: a file already present in data/acts/raw/, or
already converted to data/acts/md/, is skipped. --clean forces a full refetch.

Usage:
    python 12_select_and_download.py                 # download what is missing
    python 12_select_and_download.py --dry-run       # only report the selection
    python 12_select_and_download.py --limit 5       # smoke test
    python 12_select_and_download.py --clean         # wipe raw/ and refetch all
"""

from __future__ import annotations

import argparse
import importlib.util
import os
import random
import re
import sys
import time
import urllib.parse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
REGISTER_XLSX = BASE_DIR / "data" / "register" / "hcj_acts.xlsx"
SELECTED_XLSX = BASE_DIR / "data" / "register" / "hcj_acts_selected.xlsx"
RAW_DIR = BASE_DIR / "data" / "acts" / "raw"
MD_DIR = BASE_DIR / "data" / "acts" / "md"

OZNAKA_DISCIPLINARY = ("результати розгляду питань щодо притягнення суддів "
                       "до дисциплінарної відповідальності")
UHVALA_MARKER = "відкриття дисциплінарної справи"
SUPPORTED_EXT = {".docx", ".doc", ".rtf"}

COLUMNS = ["№", "Номер", "Вид документу", "Дата прийняття", "Назва документу",
           "doc_url", "Файл акта", "Локальний файл", "Вперше побачено"]


def _stage11():
    """Import 11_scrape_register.py (its name is not a Python identifier).

    Stage 12 reuses its register reader and, more importantly, its TLS
    handling: hcj.gov.ua omits an intermediate certificate, and that whole
    workaround should live in exactly one place.
    """
    path = Path(__file__).with_name("11_scrape_register.py")
    spec = importlib.util.spec_from_file_location("hcj_register", path)
    module = importlib.util.module_from_spec(spec)
    # Must be visible in sys.modules *before* exec: @dataclass resolves its
    # field types by looking the defining module up there.
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def norm(text: str) -> str:
    """Casefolded, whitespace-collapsed — the register's spacing is not stable."""
    return re.sub(r"\s+", " ", (text or "")).strip().casefold()


def is_selected(act) -> bool:
    if norm(act.oznaka) != OZNAKA_DISCIPLINARY:
        return False
    doctype = norm(act.doctype)
    if doctype == "рішення":
        return True
    return doctype == "ухвала" and UHVALA_MARKER in norm(act.title)


def local_name(act) -> str:
    """«1503/3дп/15-26» + «22.07.2026» → «1503_22.07.2026.docx».

    The leading number plus the date is unique across the whole selection and
    is what stages 02-04 (and the existing SQLite records) already key on.
    """
    lead = act.number.split("/", 1)[0].strip()
    ext = Path(urllib.parse.urlparse(act.file_url).path).suffix.lower()
    if ext not in SUPPORTED_EXT:
        ext = ext or ".bin"
    return f"{lead}_{act.date}{ext}"


def already_have(name: str, raw_dir: Path, md_dir: Path) -> bool:
    """True if this act has been fetched before — or already converted.

    Markdown counts: stage 13 keeps its output, so re-downloading a source file
    that has already been converted would be pure waste.
    """
    stem = Path(name).stem
    if (md_dir / f"{stem}.md").exists():
        return True
    return any((raw_dir / f"{stem}{ext}").exists()
               for ext in SUPPORTED_EXT | {Path(name).suffix.lower()})


def download(session, verify, url: str, dest: Path, retries: int = 4) -> None:
    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            resp = session.get(url, timeout=120, verify=verify)
            resp.raise_for_status()
            body = resp.content
            if not body:
                raise RuntimeError("порожня відповідь")
            # A missing file is served as a styled 200 error page, not a 404.
            if body[:512].lstrip()[:9].lower() in (b"<!doctype", b"<html"):
                raise RuntimeError("замість файлу повернуто HTML-сторінку")
            tmp = dest.with_name(dest.name + ".part")
            tmp.write_bytes(body)
            os.replace(tmp, dest)
            return
        except Exception as err:
            last_err = err
            if attempt < retries - 1:
                wait = 2 ** attempt + random.random()
                print(f"    ! {type(err).__name__}: повтор через {wait:.1f}s "
                      f"({attempt + 1}/{retries})", file=sys.stderr)
                time.sleep(wait)
    raise RuntimeError(f"{last_err}")


def export_selection(acts: list, names: dict, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Дисциплінарні акти"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center",
                                   wrap_text=True)

    for i, act in enumerate(acts, start=1):
        ws.append([i, act.number, act.doctype, act.date, act.title,
                   act.doc_url, act.file_url, names[act.key], act.first_seen])

    for col, width in enumerate([6, 18, 16, 15, 90, 34, 34, 28, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    wb.save(tmp)
    os.replace(tmp, path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Відбір дисциплінарних актів з реєстру ВРП і завантаження їх файлів",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--register", default=str(REGISTER_XLSX),
                        help="Таблиця реєстру (результат стадії 00)")
    parser.add_argument("--out", default=str(SELECTED_XLSX),
                        help="Куди записати відібрані рядки")
    parser.add_argument("--raw-dir", default=str(RAW_DIR),
                        help="Куди складати завантажені файли актів")
    parser.add_argument("--md-dir", default=str(MD_DIR),
                        help="Де стадія 13 тримає Markdown (уже конвертовані "
                             "акти повторно не завантажуються)")
    parser.add_argument("--clean", action="store_true",
                        help="Видалити всі раніше завантажені файли й "
                             "завантажити все наново")
    parser.add_argument("--dry-run", action="store_true",
                        help="Лише показати відбір, нічого не завантажувати")
    parser.add_argument("--limit", type=int, default=None,
                        help="Завантажити не більше N файлів (для перевірки)")
    parser.add_argument("--delay", type=float, default=0.7,
                        help="Пауза між завантаженнями, с")
    parser.add_argument("--insecure", action="store_true",
                        help="Не перевіряти TLS-сертифікат взагалі")
    parser.add_argument("--strict", action="store_true",
                        help="Перервати роботу, якщо перевірити сертифікат не вдалося")
    args = parser.parse_args(argv)

    if args.insecure and args.strict:
        parser.error("--insecure та --strict взаємно виключні")

    stage11 = _stage11()
    register_path = Path(args.register)
    if not register_path.exists():
        raise SystemExit(f"ПОМИЛКА: реєстр не знайдено: {register_path}\n"
                         "Спершу запустіть 11_scrape_register.py.")

    acts = stage11.load_register(register_path)
    selected = [act for act in acts if is_selected(act)]
    by_type: dict[str, int] = {}
    for act in selected:
        by_type[act.doctype] = by_type.get(act.doctype, 0) + 1
    print(f"Реєстр: {len(acts)} актів → відібрано {len(selected)}"
          f" ({', '.join(f'{k}: {v}' for k, v in sorted(by_type.items()))})")

    names = {act.key: local_name(act) for act in selected}
    clashes = len(selected) - len(set(names.values()))
    if clashes:
        print(f"УВАГА: {clashes} актів дають однакове ім'я файлу — частину буде "
              f"перезаписано.", file=sys.stderr)

    export_selection(selected, names, Path(args.out))
    print(f"Відбір збережено → {args.out}")

    raw_dir = Path(args.raw_dir)
    md_dir = Path(args.md_dir)
    raw_dir.mkdir(parents=True, exist_ok=True)

    if args.clean:
        removed = 0
        for f in raw_dir.iterdir():
            if f.is_file():
                f.unlink()
                removed += 1
        print(f"--clean: видалено {removed} раніше завантажених файлів")

    if args.dry_run:
        missing = sum(1 for act in selected
                      if act.file_url
                      and not already_have(names[act.key], raw_dir, md_dir))
        print(f"--dry-run: бракує {missing} файлів, нічого не завантажено")
        return 0

    # --clean means "fetch everything again", so it also ignores the Markdown
    # that would otherwise mark a document as already processed.
    todo = [act for act in selected if act.file_url
            and (args.clean or not already_have(names[act.key], raw_dir, md_dir))]
    if args.limit:
        todo = todo[:args.limit]
    print(f"Уже на диску: {len(selected) - len(todo)}; до завантаження: {len(todo)}")
    if not todo:
        return 0

    session, verify = stage11.make_session(args.insecure, args.strict)
    ok, failed = 0, 0
    for i, act in enumerate(todo, start=1):
        name = names[act.key]
        print(f"[{i}/{len(todo)}] {name} ...", flush=True)
        try:
            download(session, verify, act.file_url, raw_dir / name)
            ok += 1
        except Exception as err:
            failed += 1
            print(f"   ❌ {name}: {err}", file=sys.stderr)
        if i < len(todo):
            time.sleep(args.delay)

    print(f"\nГотово: {ok} завантажено, {failed} з помилкою → {raw_dir}")
    # Individual failures are not fatal: the file stays missing and the next
    # run retries it. Everything failing means the site is unreachable.
    if ok == 0 and failed:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
