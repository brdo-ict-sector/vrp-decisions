#!/usr/bin/env python3
"""
Scraper for the register of acts of the High Council of Justice of Ukraine
(Вища рада правосуддя) — https://hcj.gov.ua/acts

Collects the table columns:
    №, Номер, Вид документу, Дата прийняття, Назва документу, Ознака до документа
and exports them to an .xlsx file.

The listing is a Drupal View with an exposed GET filter form, 20 rows per page:
    /acts?type=<tid|All>
         &number=<str>
         &date[value][date]=<dd.mm.yyyy>        -> "Дата прийняття - з"
         &date_filter_1[value][date]=<dd.mm.yyyy> -> "Дата прийняття - по"
         &field_title_value=<str>
         &text=<str>
         &field_oznaka_tid_i18n=<tid|All>
         &page=<0-based page index>

This is stage 11 of the pipeline: it builds the *index* of the register — one
row per act, including the link to the act's page (`doc_url`) and to its source
file (`Файл акта`).  That index is what tells stage 12 which documents exist and
where to fetch them from.

The register is kept as ONE growing table, data/register/hcj_acts.xlsx.  A run
never replaces it: freshly scraped rows are merged into whatever is already
there, keyed on `doc_url` (the only truly unique key — the register legitimately
contains several acts sharing one number, date and title).  Rows the merge has
not seen before are stamped in the «Вперше побачено» column, which gives a free
change log of what each daily run added.

By default the run scrapes only the last --since-days days, which is what the
nightly job needs: cheap, and wide enough to pick up acts the HCJ publishes with
a delay under an earlier date.  Use --full for a complete (re)build.

Usage:
    python 11_scrape_register.py                    # last 30 days → hcj_acts.xlsx
    python 11_scrape_register.py --since-days 90
    python 11_scrape_register.py --full --date-from 01.01.2025
    python 11_scrape_register.py --seed hcj_acts_2025.xlsx hcj_acts_2026.xlsx
    python 11_scrape_register.py --oznaka disciplinary -o only_disc.xlsx
"""

from __future__ import annotations

import argparse
import os
import random
import re
import ssl
import sys
import tempfile
import time
import urllib.parse
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE_URL = "https://hcj.gov.ua/acts"
ROWS_PER_PAGE = 20
# Same convention as the other pipeline stages: paths resolve from the repo
# root, not the current working directory.
BASE_DIR = Path(__file__).resolve().parent.parent
REGISTER_DIR = BASE_DIR / "data" / "register"
REGISTER_XLSX = "hcj_acts.xlsx"  # the one growing table, inside REGISTER_DIR
DATE_FMT = "%d.%m.%Y"
# The intermediate CA certificate that hcj.gov.ua omits from its TLS chain.
INTERMEDIATE_PEM = Path(__file__).with_name("hcj_intermediate.pem")
USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# "Вид документу" — select#edit-type values on the site
DOC_TYPES = {
    "Рішення": "24",
    "Постанова": "25",
    "Ухвала": "26",
}

# "Ознака до документа" — select#edit-field-oznaka-tid-i18n values on the site
OZNAKA = {
    "disciplinary": ("91", "Результати розгляду питань щодо притягнення суддів "
                           "до дисциплінарної відповідальності"),
    "interference": ("92", "Результати розгляду повідомлень суддів про втручання "
                           "в діяльність"),
    "suspension": ("93", "Відсторонення від здійснення правосуддя"),
    "dismissal-general": ("94", "Звільнення суддів за загальними обставинами"),
    "dismissal-special": ("95", "Звільнення суддів за особливими обставинами"),
}

COLUMNS = [
    "№",
    "Номер",
    "Вид документу",
    "Дата прийняття",
    "Назва документу",
    "Ознака до документа",
    "Примітка",
    "doc_url",
    "Файл акта",
    "Вперше побачено",
]


@dataclass
class Act:
    index: str = ""
    number: str = ""
    doctype: str = ""
    date: str = ""
    title: str = ""
    oznaka: str = ""
    notes: str = ""
    doc_url: str = ""
    file_url: str = ""
    first_seen: str = ""

    @property
    def key(self) -> tuple | str:
        """Merge/dedup key. The node URL is unique per act; several acts can
        otherwise share one number, date and title (separate files)."""
        return self.doc_url or (self.number, self.date, self.title, self.file_url)

    def as_row(self) -> list:
        return [
            int(self.index) if self.index.isdigit() else self.index,
            self.number,
            self.doctype,
            self.date,
            self.title,
            self.oznaka,
            self.notes,
            self.doc_url,
            self.file_url,
            self.first_seen,
        ]


@dataclass
class Filters:
    date_from: str = ""
    date_to: str = ""
    doc_type: str = "All"
    number: str = ""
    title: str = ""
    text: str = ""
    oznaka: str = "All"

    def query(self, page: int = 0) -> dict:
        params = {
            # The view's default ordering (date desc) is not stable across
            # requests: many acts share one date, so rows shuffle between page
            # loads — that both duplicates and silently skips records.  Sorting
            # by the unique document number makes pagination deterministic.
            "order": "field_number",
            "sort": "asc",
            "type": self.doc_type,
            "number": self.number,
            "date[value][date]": self.date_from,
            "date_filter_1[value][date]": self.date_to,
            "field_title_value": self.title,
            "text": self.text,
            "field_oznaka_tid_i18n": self.oznaka,
        }
        if page:
            params["page"] = str(page)
        return params


def make_session(insecure: bool = False,
                 strict: bool = False) -> tuple[requests.Session, bool | str]:
    """Build a session; returns (session, value for requests' `verify=`).

    That value is True (chain verified normally), a path to a CA bundle
    (verified using the intermediate the server fails to send), or False
    (verification disabled — only as a last resort, or on --insecure).
    """
    session = requests.Session()
    session.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "uk-UA,uk;q=0.9,en;q=0.8",
    })

    if insecure:
        silence_tls_warnings()
        return session, False

    # 1. Normal verification — works if the server ever fixes its chain.
    try:
        session.get(BASE_URL, timeout=30)
        return session, True
    except requests.exceptions.SSLError:
        pass

    # 2. Verification against certifi's roots + the bundled intermediate.
    bundle = build_ca_bundle()
    if bundle:
        try:
            session.get(BASE_URL, timeout=30, verify=bundle)
            return session, bundle
        except requests.exceptions.SSLError as err:
            print(f"  ! з локальним CA-бандлом: {str(err)[:120]}", file=sys.stderr)

    # 3. Certificates rotate — try to discover the current intermediate from
    #    the leaf's Authority Information Access extension and retry.
    fetched = fetch_intermediate_via_aia()
    if fetched:
        bundle = build_ca_bundle(extra=fetched)
        try:
            session.get(BASE_URL, timeout=30, verify=bundle)
            print("Проміжний сертифікат отримано через AIA — перевірку TLS "
                  "відновлено.", file=sys.stderr)
            return session, bundle
        except requests.exceptions.SSLError:
            pass

    if strict:
        raise SystemExit(
            "ПОМИЛКА: не вдалося перевірити TLS-сертифікат hcj.gov.ua, а вказано "
            "--strict. Запустіть без --strict або оновіть "
            f"{INTERMEDIATE_PEM.name}."
        )

    print("УВАГА: не вдалося побудувати довірений ланцюжок сертифікатів — "
          "перевірку TLS вимкнено.", file=sys.stderr)
    silence_tls_warnings()
    return session, False


def silence_tls_warnings() -> None:
    requests.packages.urllib3.disable_warnings(
        requests.packages.urllib3.exceptions.InsecureRequestWarning
    )


def build_ca_bundle(extra: bytes | None = None) -> str | None:
    """Write certifi's roots + our intermediate(s) to a temp PEM; return path.

    hcj.gov.ua sends only its leaf certificate, omitting the intermediate CA
    that links it to a trusted root, so stock verification cannot complete the
    chain.  Supplying that intermediate ourselves restores real verification —
    the certificate is still checked against a public root, we merely provide
    the link the server should have sent.
    """
    parts: list[bytes] = []
    try:
        import certifi
        parts.append(Path(certifi.where()).read_bytes())
    except Exception:
        return None

    if INTERMEDIATE_PEM.exists():
        parts.append(INTERMEDIATE_PEM.read_bytes())
    if extra:
        parts.append(extra)
    if len(parts) < 2:
        return None

    bundle = Path(tempfile.gettempdir()) / "hcj_ca_bundle.pem"
    bundle.write_bytes(b"\n".join(parts))
    return str(bundle)


def fetch_intermediate_via_aia() -> bytes | None:
    """Download the issuing CA certificate named in the leaf's AIA extension."""
    try:
        from cryptography import x509
        from cryptography.hazmat.primitives.serialization import Encoding
    except ImportError:
        return None

    host = urllib.parse.urlparse(BASE_URL).hostname or ""
    try:
        leaf_pem = ssl.get_server_certificate((host, 443))
        leaf = x509.load_pem_x509_certificate(leaf_pem.encode())
        aia = leaf.extensions.get_extension_for_class(
            x509.AuthorityInformationAccess).value
        urls = [
            desc.access_location.value
            for desc in aia
            if desc.access_method == x509.oid.AuthorityInformationAccessOID.CA_ISSUERS
        ]
    except Exception:
        return None

    for url in urls:
        try:
            raw = requests.get(url, timeout=30).content
            try:
                cert = x509.load_der_x509_certificate(raw)
            except ValueError:
                cert = x509.load_pem_x509_certificate(raw)
            return cert.public_bytes(Encoding.PEM)
        except Exception:
            continue
    return None


def fetch(session: requests.Session, verify: bool | str, params: dict,
          retries: int = 4) -> str:
    url = f"{BASE_URL}?{urllib.parse.urlencode(params)}"
    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            resp = session.get(url, timeout=60, verify=verify)
            resp.raise_for_status()
            resp.encoding = resp.encoding or "utf-8"
            return resp.text
        except Exception as err:  # network hiccups / 5xx / timeouts
            last_err = err
            wait = 2 ** attempt + random.random()
            print(f"  ! {type(err).__name__}: повтор через {wait:.1f}s "
                  f"({attempt + 1}/{retries})", file=sys.stderr)
            time.sleep(wait)
    raise RuntimeError(f"Не вдалося завантажити {url}: {last_err}")


def cell_text(td) -> str:
    return re.sub(r"\s+", " ", td.get_text(" ", strip=True)).strip()


def cell_link(td) -> str:
    a = td.find("a", href=True)
    return urllib.parse.urljoin(BASE_URL, a["href"]) if a else ""


def parse_page(html: str) -> tuple[list[Act], int | None]:
    """Return (acts on this page, total number of pages if discoverable)."""
    soup = BeautifulSoup(html, "lxml")

    acts: list[Act] = []
    table = soup.select_one("table.views-table")
    if table:
        for tr in table.select("tbody > tr"):
            tds = {}
            for td in tr.find_all("td", recursive=False):
                for cls in td.get("class", []):
                    if cls.startswith("views-field-") and cls != "views-field":
                        tds[cls] = td
            if not tds:
                continue
            get = lambda k: tds.get(f"views-field-{k}")  # noqa: E731
            num_td = get("field-number")
            title_td = get("field-title")
            act = Act(
                index=cell_text(get("counter")) if get("counter") else "",
                number=cell_text(num_td) if num_td else "",
                doctype=cell_text(get("field-doctype")) if get("field-doctype") else "",
                date=cell_text(get("field-date-iso")) if get("field-date-iso") else "",
                title=cell_text(title_td) if title_td else "",
                oznaka=cell_text(get("field-oznaka")) if get("field-oznaka") else "",
                notes=cell_text(get("field-notes")) if get("field-notes") else "",
                doc_url=cell_link(title_td) if title_td else "",
                file_url=cell_link(num_td) if num_td else "",
            )
            if act.number or act.title:
                acts.append(act)

    total_pages = None
    last = soup.select_one("ul.pagination li.pager-last a[href]")
    if not last:
        # single page, or we are already on the last one
        nums = [
            int(m.group(1))
            for a in soup.select("ul.pagination a[href]")
            if (m := re.search(r"[?&]page=(\d+)", a["href"]))
        ]
        total_pages = (max(nums) + 1) if nums else 1
    else:
        m = re.search(r"[?&]page=(\d+)", last["href"])
        if m:
            total_pages = int(m.group(1)) + 1
    return acts, total_pages


def scrape(filters: Filters, delay: float, max_pages: int | None,
           insecure: bool, strict: bool = False) -> list[Act]:
    session, verify = make_session(insecure, strict)

    print(f"Фільтр: з {filters.date_from or '—'} по {filters.date_to or '—'}, "
          f"вид={filters.doc_type}, ознака={filters.oznaka}")

    acts: list[Act] = []
    seen: set = set()
    page = 0
    total_pages: int | None = None

    while True:
        html = fetch(session, verify, filters.query(page))
        page_acts, discovered = parse_page(html)
        if total_pages is None and discovered:
            total_pages = discovered
            if max_pages:
                total_pages = min(total_pages, max_pages)
            print(f"Знайдено сторінок: {discovered}"
                  + (f" (обробляємо {total_pages})" if total_pages != discovered else "")
                  + f", приблизно записів: ~{discovered * ROWS_PER_PAGE}")

        new = 0
        for act in page_acts:
            if act.key in seen:
                continue
            seen.add(act.key)
            acts.append(act)
            new += 1

        print(f"  стор. {page + 1}"
              + (f"/{total_pages}" if total_pages else "")
              + f": {len(page_acts)} рядків (+{new}), усього {len(acts)}")

        if not page_acts or new == 0:
            break
        page += 1
        if total_pages is not None and page >= total_pages:
            break
        if max_pages and page >= max_pages:
            break
        time.sleep(delay)

    expected = (total_pages or 0) * ROWS_PER_PAGE
    if total_pages and not max_pages and len(acts) < expected - ROWS_PER_PAGE:
        print(f"УВАГА: зібрано {len(acts)} записів, а сторінок {total_pages} "
              f"(очікувалось до ~{expected}). Можливо, частину даних пропущено.",
              file=sys.stderr)

    return sort_acts(acts)


def sort_acts(acts: list[Act]) -> list[Act]:
    """Newest first, then by document number — and renumber the '№' column.

    We paginate sorted by number (for stability), so the site's own row counter
    does not correspond to this ordering; it is regenerated as 1..N.
    """
    def key(act: Act):
        try:
            day, month, year = (int(p) for p in act.date.split("."))
            date_key = (year, month, day)
        except ValueError:
            date_key = (0, 0, 0)
        m = re.match(r"(\d+)", act.number)
        return (date_key, int(m.group(1)) if m else 0, act.number)

    acts = sorted(acts, key=key, reverse=True)
    for i, act in enumerate(acts, start=1):
        act.index = str(i)
    return acts


def load_register(path: Path) -> list[Act]:
    """Read an existing register table. Missing file → empty list.

    Column *names* are honoured rather than positions, so tables written by
    earlier versions (without «Вперше побачено») still load.
    """
    if not path.exists():
        return []

    ws = load_workbook(path, read_only=True).active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
    except StopIteration:
        return []

    fields = {
        "Номер": "number",
        "Вид документу": "doctype",
        "Дата прийняття": "date",
        "Назва документу": "title",
        "Ознака до документа": "oznaka",
        "Примітка": "notes",
        "doc_url": "doc_url",
        "Файл акта": "file_url",
        "Вперше побачено": "first_seen",
    }
    idx = {fields[name]: i for i, name in enumerate(header) if name in fields}
    if "doc_url" not in idx:
        raise SystemExit(f"ПОМИЛКА: у {path} немає колонки doc_url — це не реєстр актів.")

    acts = []
    for row in rows:
        def val(field: str) -> str:
            i = idx.get(field)
            return "" if i is None or i >= len(row) or row[i] is None else str(row[i]).strip()

        act = Act(**{field: val(field) for field in idx})
        if act.number or act.title:
            acts.append(act)
    return acts


def merge_acts(existing: list[Act], incoming: list[Act], stamp: str) -> tuple[list[Act], int]:
    """Merge `incoming` into `existing`; return (merged, number of new acts).

    Acts already in the table keep their «Вперше побачено» stamp and are
    refreshed from the site (titles and file links do get corrected there).
    """
    by_key = {act.key: act for act in existing}
    new = 0
    for act in incoming:
        old = by_key.get(act.key)
        if old is None:
            act.first_seen = stamp
            by_key[act.key] = act
            new += 1
        else:
            # Keep the original stamp verbatim — including the empty one that
            # marks seeded history. Re-reading an act does not make it new.
            act.first_seen = old.first_seen
            by_key[act.key] = act
    return list(by_key.values()), new


def export_xlsx(acts: list[Act], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Акти ВРП"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center",
                                   wrap_text=True)

    for act in acts:
        ws.append(act.as_row())

    widths = [6, 18, 16, 15, 90, 45, 25, 34, 34, 18]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wrap = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    # Write beside the target, then rename: a crash (or a nightly run killed
    # mid-save) must never leave the register truncated.
    target = Path(path)
    tmp = target.with_name(target.name + ".tmp")
    wb.save(tmp)
    os.replace(tmp, target)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Вивантаження реєстру актів ВРП (hcj.gov.ua/acts) у xlsx",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--date-from", default="",
                        help="'Дата прийняття - з' у форматі дд.мм.рррр "
                             "(типово: сьогодні мінус --since-days)")
    parser.add_argument("--date-to", default="",
                        help="'Дата прийняття - по' у форматі дд.мм.рррр")
    parser.add_argument("--since-days", type=int, default=30,
                        help="Скільки останніх днів переглядати, якщо "
                             "--date-from не задано")
    parser.add_argument("--full", action="store_true",
                        help="Повний перегляд реєстру: з --date-from (типово "
                             "01.01.2025) до сьогодні")
    parser.add_argument("--seed", nargs="+", metavar="XLSX", default=[],
                        help="Влити наявні xlsx-таблиці в реєстр без скрейпінгу")
    parser.add_argument("--type", dest="doc_type", default="All",
                        choices=["All", *DOC_TYPES], help="Вид документу")
    parser.add_argument("--oznaka", default="All",
                        choices=["All", *OZNAKA], help="Ознака до документа")
    parser.add_argument("--number", default="", help="Фільтр за номером")
    parser.add_argument("--title", default="", help="Пошук по назві документу")
    parser.add_argument("--text", default="", help="Пошук у тексті документу")
    parser.add_argument("-o", "--out", default=REGISTER_XLSX,
                        help="Таблиця реєстру: у неї вливається результат "
                             "(відносні шляхи — у data/register/)")
    parser.add_argument("--delay", type=float, default=0.7,
                        help="Пауза між запитами, с")
    parser.add_argument("--max-pages", type=int, default=None,
                        help="Обмежити кількість сторінок (для тесту)")
    parser.add_argument("--insecure", action="store_true",
                        help="Не перевіряти TLS-сертифікат взагалі")
    parser.add_argument("--strict", action="store_true",
                        help="Перервати роботу, якщо перевірити сертифікат "
                             "не вдалося (без переходу в незахищений режим)")
    args = parser.parse_args(argv)

    date_re = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
    for label, value in (("--date-from", args.date_from), ("--date-to", args.date_to)):
        if value and not date_re.match(value):
            parser.error(f"{label} має бути у форматі дд.мм.рррр, отримано {value!r}")
    if args.insecure and args.strict:
        parser.error("--insecure та --strict взаємно виключні")

    today = date.today()
    if args.date_from:
        date_from = args.date_from
    elif args.full:
        date_from = "01.01.2025"
    else:
        # Acts are often published days after they were adopted, so a daily run
        # has to re-read a window, not just yesterday.
        date_from = (today - timedelta(days=args.since_days)).strftime(DATE_FMT)

    # A bare filename lands in data/register/; an explicit path is honoured.
    out = Path(args.out)
    if not out.is_absolute() and out.parent == Path("."):
        out = REGISTER_DIR / out

    register = load_register(out)
    print(f"Реєстр: {out} — {len(register)} записів на початок")
    stamp = today.strftime(DATE_FMT)

    if args.seed:
        incoming: list[Act] = []
        for src in args.seed:
            src_path = Path(src)
            if not src_path.is_absolute() and src_path.parent == Path("."):
                src_path = REGISTER_DIR / src_path
            rows = load_register(src_path)
            print(f"  + {src_path.name}: {len(rows)} записів")
            incoming.extend(rows)
        # Imported history is not something *this* run discovered; leave the
        # stamp empty so «Вперше побачено» stays an honest change log.
        merged, added = merge_acts(register, incoming, stamp="")
    else:
        filters = Filters(
            date_from=date_from,
            date_to=args.date_to,
            doc_type=DOC_TYPES.get(args.doc_type, "All"),
            number=args.number,
            title=args.title,
            text=args.text,
            oznaka=OZNAKA[args.oznaka][0] if args.oznaka in OZNAKA else "All",
        )
        scraped = scrape(filters, args.delay, args.max_pages, args.insecure, args.strict)
        if not scraped:
            # An empty window is normal (holidays); an empty *full* scrape is not.
            if not register or args.full:
                print("Жодного запису не знайдено — реєстр не змінено.", file=sys.stderr)
                return 1
            print("\nНових актів немає — реєстр без змін.")
            return 0
        merged, added = merge_acts(register, scraped, stamp)

    merged = sort_acts(merged)
    out.parent.mkdir(parents=True, exist_ok=True)
    export_xlsx(merged, str(out))
    print(f"\nГотово: +{added} нових, усього {len(merged)} записів → {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
